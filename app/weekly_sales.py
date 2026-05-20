from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


WEEKLY_SALES_RE = re.compile(r"^Vta\.Sem\.(\d+)$", re.IGNORECASE)
FILENAME_WEEK_RE = re.compile(r"\b(?:semana|sem)\s*(\d{1,2})\b", re.IGNORECASE)


@dataclass(slots=True)
class WeeklyAgencySale:
    source_file: str
    source_sheet: str
    source_row: int
    week: int
    lotos_code: str | None
    master_code: str | None
    previous_lotos_code: str | None
    agent_name: str | None
    rut: str | None
    address: str | None
    comuna: str | None
    region_number: str | None
    rubro: str | None
    executive: str | None
    admission_date: str | None
    commercial_status: str | None
    operational_status: str | None
    top_segment: str | None
    coverage: str | None
    sales_status: str | None
    weekly_sales: float
    average_sales_2019: float | None
    difference_vs_2019: float | None
    latitude: float | None
    longitude: float | None
    territory: str | None
    closed_date: str | None

    @property
    def is_selling(self) -> bool:
        return self.weekly_sales > 0

    @property
    def is_closed(self) -> bool:
        closed_values = {"baja", "direccion baja", "dirección baja"}
        return _normalized(self.commercial_status) in closed_values or _normalized(self.sales_status) in closed_values


@dataclass(slots=True)
class WeeklyWorkbookParseResult:
    rows: list[WeeklyAgencySale]
    skipped_sheets: list[str]


def parse_weekly_workbooks(paths: Iterable[str | Path]) -> list[WeeklyAgencySale]:
    rows_by_key: dict[tuple[str, int], WeeklyAgencySale] = {}
    for path in paths:
        for row in parse_weekly_workbook(path).rows:
            if not row.lotos_code:
                continue
            rows_by_key[(row.lotos_code, row.week)] = row
    return sorted(rows_by_key.values(), key=lambda row: (row.week, row.lotos_code or ""))


def parse_weekly_workbook(path: str | Path) -> WeeklyWorkbookParseResult:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[WeeklyAgencySale] = []
    skipped_sheets: list[str] = []
    filename_week = _week_from_filename(workbook_path.name)

    for sheet in workbook.worksheets:
        header_row, headers, sales_headers = _find_header(sheet)
        if header_row is None or headers is None or not sales_headers:
            skipped_sheets.append(sheet.title)
            continue

        normalized_headers = [_normalized_header(header) for header in headers]
        for row_index, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if _is_empty_row(row):
                continue
            row_map = dict(zip(normalized_headers, row))
            lotos_code = _code(row_map.get("lotos"))
            if not lotos_code:
                continue
            for header_week, sales_header in sales_headers:
                week = (filename_week or header_week) if len(sales_headers) == 1 else header_week
                weekly_sales = _number(row_map.get(sales_header)) or 0.0
                rows.append(
                    WeeklyAgencySale(
                        source_file=workbook_path.name,
                        source_sheet=sheet.title,
                        source_row=row_index,
                        week=week,
                        lotos_code=lotos_code,
                        master_code=_code(row_map.get("master")),
                        previous_lotos_code=_code(row_map.get("loto anterior")),
                        agent_name=_text(row_map.get("nombre agente")),
                        rut=_code(row_map.get("rut.")),
                        address=_text(row_map.get("direccion")),
                        comuna=_text(row_map.get("comuna")),
                        region_number=_code(row_map.get("reg.")),
                        rubro=_text(row_map.get("rubro")),
                        executive=_text(row_map.get("ejec./ coord.")),
                        admission_date=_date(row_map.get("ingreso")),
                        commercial_status=_text(row_map.get("est. com.")),
                        operational_status=_text(row_map.get("status")),
                        top_segment=_text(row_map.get("top")),
                        coverage=_text(row_map.get("cobertura")),
                        sales_status=_text(row_map.get("vta.")),
                        weekly_sales=weekly_sales,
                        average_sales_2019=_number(row_map.get("vta prom. 2019")),
                        difference_vs_2019=_number(row_map.get("diferencia")),
                        latitude=_number(row_map.get("latitud")),
                        longitude=_number(row_map.get("longitud")),
                        territory=_text(row_map.get("ubicacion")),
                        closed_date=_date(row_map.get("baja")),
                    )
                )

    return WeeklyWorkbookParseResult(rows=rows, skipped_sheets=skipped_sheets)


def _find_header(sheet) -> tuple[int | None, list[object] | None, list[tuple[int, str]]]:
    for row_index in range(1, min(sheet.max_row, 15) + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
        normalized_headers = [_normalized_header(value) for value in values]
        if "lotos" not in normalized_headers:
            continue
        sales_headers: list[tuple[int, str]] = []
        for header in normalized_headers:
            match = WEEKLY_SALES_RE.match(header)
            if match:
                sales_headers.append((int(match.group(1)), header))
        if sales_headers:
            return row_index, values, sorted(sales_headers, key=lambda item: item[0])
    return None, None, []


def _week_from_filename(filename: str) -> int | None:
    match = FILENAME_WEEK_RE.search(filename)
    if not match:
        return None
    return int(match.group(1))


def _is_empty_row(row: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _normalized_header(value: object) -> str:
    text = _text(value) or ""
    text = text.replace("Dirección", "Direccion").replace("Ubicación", "Ubicacion")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def _normalized(value: object) -> str:
    text = _text(value) or ""
    text = text.replace("Dirección", "Direccion")
    return re.sub(r"\s+", " ", text).strip().lower()


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _code(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _text(value)


def _number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _text(value)
