from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class AgencyPrizeSummary:
    lotos_code: str
    agent_name: str | None
    gross_total: int
    net_total: int
    subgames_count: int
    top_subgames: list[dict[str, int | str]]


def parse_agency_prize_workbook(path: str | Path) -> list[AgencyPrizeSummary]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["Informe 1"] if "Informe 1" in workbook.sheetnames else workbook.worksheets[0]

    by_agency: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        lotos_code = _code(row[1] if len(row) > 1 else None)
        if not lotos_code:
            continue
        agent_name = _text(row[2] if len(row) > 2 else None)
        subgame = _text(row[3] if len(row) > 3 else None)
        gross = int(_number(row[4] if len(row) > 4 else None) or 0)
        net = int(_number(row[5] if len(row) > 5 else None) or 0)
        if gross <= 0 and net <= 0:
            continue
        item = by_agency.setdefault(
            lotos_code,
            {
                "agent_name": agent_name,
                "gross_total": 0,
                "net_total": 0,
                "subgames": {},
            },
        )
        item["gross_total"] += gross
        item["net_total"] += net
        if subgame:
            game = item["subgames"].setdefault(subgame, {"gross_total": 0, "net_total": 0})
            game["gross_total"] += gross
            game["net_total"] += net

    rows = []
    for lotos_code, item in sorted(by_agency.items()):
        top_subgames = sorted(
            (
                {"subgame": subgame, "gross_total": values["gross_total"], "net_total": values["net_total"]}
                for subgame, values in item["subgames"].items()
            ),
            key=lambda subgame: subgame["gross_total"],
            reverse=True,
        )
        rows.append(
            AgencyPrizeSummary(
                lotos_code=lotos_code,
                agent_name=item["agent_name"],
                gross_total=item["gross_total"],
                net_total=item["net_total"],
                subgames_count=len(top_subgames),
                top_subgames=top_subgames[:5],
            )
        )
    return rows


def find_agency_prize_workbook(input_dir: str | Path) -> Path | None:
    matches = sorted(Path(input_dir).glob("Premios Vendidos por Lotos*.xlsx"))
    return matches[-1] if matches else None


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _code(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _text(value)


def _number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None
