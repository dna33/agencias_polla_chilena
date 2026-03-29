from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.geo import parse_coordinates
from app.models import Agency
from app.repository import AgencyRepository
from app.scheduler import build_schedule

TARGET_SHEET = "Informe 1"
HEADER_ROW_INDEX = 4
DATA_START_ROW_INDEX = 5


@dataclass(slots=True)
class ImportReport:
    total_rows: int = 0
    imported_rows: int = 0
    searchable_rows: int = 0
    discarded_rows: int = 0
    invalid_coordinates_rows: list[int] = field(default_factory=list)
    invalid_schedule_rows: list[int] = field(default_factory=list)
    ineligible_status_rows: list[int] = field(default_factory=list)


class ExcelAgencyImporter:
    def __init__(self, repository: AgencyRepository | None = None) -> None:
        self.repository = repository or AgencyRepository()

    def import_file(self, file_path: str) -> ImportReport:
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook[TARGET_SHEET]
        headers = self._read_headers(sheet)
        agencies: list[Agency] = []
        report = ImportReport()

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=DATA_START_ROW_INDEX, values_only=True),
            start=DATA_START_ROW_INDEX,
        ):
            if self._is_empty_row(row):
                continue
            report.total_rows += 1
            row_map = dict(zip(headers, row))
            agency = self._normalize_row(row_map, row_index)
            agencies.append(agency)
            report.imported_rows += 1

            if agency.latitude is None or agency.longitude is None:
                report.invalid_coordinates_rows.append(row_index)
            if any(
                error.startswith(
                    (
                        "monday:invalid_schedule",
                        "tuesday:invalid_schedule",
                        "wednesday:invalid_schedule",
                        "thursday:invalid_schedule",
                        "friday:invalid_schedule",
                        "saturday:invalid_schedule",
                        "sunday:invalid_schedule",
                    )
                )
                for error in agency.data_quality_errors
            ):
                report.invalid_schedule_rows.append(row_index)
            if commercial_or_agent_status_ineligible(agency):
                report.ineligible_status_rows.append(row_index)
            if not agency.is_active_for_search:
                report.discarded_rows += 1
            else:
                report.searchable_rows += 1

        self.repository.replace_all(agencies)
        return report

    def _read_headers(self, sheet) -> list[str]:
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[HEADER_ROW_INDEX]]
        return headers

    def _normalize_row(self, row: dict[str, object], row_index: int) -> Agency:
        schedule_json, schedule_raw_json, schedule_errors = build_schedule(row)
        latitude, longitude, coordinate_errors = parse_coordinates(_text(row.get("Coord. Geográficas")))
        commercial_status = _text(row.get("Estado Comercial"))
        agent_status = _text(row.get("Estado Agente"))
        is_active_for_search = is_search_eligible(
            commercial_status=commercial_status,
            agent_status=agent_status,
            latitude=latitude,
            longitude=longitude,
            schedule_errors=schedule_errors,
        )
        raw_row_hash = hashlib.sha256(repr(sorted((key, _text(value)) for key, value in row.items())).encode("utf-8")).hexdigest()
        status_change_date = _coerce_excel_date(row.get("Fecha Cambio Estado"))
        data_quality_errors = schedule_errors + coordinate_errors
        if commercial_status != "Activo":
            data_quality_errors.append(f"ineligible_commercial_status:{commercial_status}")
        if agent_status != "Active":
            data_quality_errors.append(f"ineligible_agent_status:{agent_status}")

        return Agency(
            lotos_code=_text(row.get("Lotos")),
            master_code=_text(row.get("Master")),
            raspe_code=_text(row.get("Pto.Cons.Raspe")),
            agent_name=_text(row.get("Nombre Agente")),
            rut=_text(row.get("Rut")),
            address=_text(row.get("Dirección")),
            comuna=_text(row.get("Comuna")),
            region_number=_text(row.get("Nro. Región")),
            rubro=_text(row.get("Rubro")),
            legal_representative=_text(row.get("Representante Legal")),
            phone_local=_text(row.get("Teléfono Local")),
            phone_1=_text(row.get("Teléfono 1")),
            phone_2=_text(row.get("Teléfono 2")),
            email=_text(row.get("E-Mail")),
            contact_name=_text(row.get("Contacto Local")),
            observation=_text(row.get("Observación")),
            commercial_status=commercial_status,
            agent_status=agent_status,
            status_change_date=status_change_date,
            latitude=latitude,
            longitude=longitude,
            raw_coordinates=_text(row.get("Coord. Geográficas")),
            schedule_json=schedule_json,
            schedule_raw_json=schedule_raw_json,
            data_quality_errors=data_quality_errors,
            raw_row_hash=raw_row_hash,
            is_active_for_search=is_active_for_search,
        )

    def _is_empty_row(self, row: tuple[object, ...]) -> bool:
        return all(cell is None or str(cell).strip() == "" for cell in row)


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _coerce_excel_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _text(value)


def is_search_eligible(
    commercial_status: str | None,
    agent_status: str | None,
    latitude: float | None,
    longitude: float | None,
    schedule_errors: list[str],
) -> bool:
    return (
        commercial_status == "Activo"
        and agent_status == "Active"
        and latitude is not None
        and longitude is not None
        and not schedule_errors
    )


def commercial_or_agent_status_ineligible(agency: Agency) -> bool:
    return agency.commercial_status != "Activo" or agency.agent_status != "Active"
