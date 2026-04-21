from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


COMUNA_SHEET = "LOTO_ Comuna"
POINT_OF_SALE_SHEET = "LOTO_ PtoVta"
HEADER_ROW = 8
DATA_START_ROW = 9


@dataclass(slots=True)
class WeeklyZoneSale:
    zone: str
    week: int
    week_label: str
    sales: int
    communes: int


def parse_weekly_zone_evolution(path: str | Path) -> list[WeeklyZoneSale]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if COMUNA_SHEET not in workbook.sheetnames or POINT_OF_SALE_SHEET not in workbook.sheetnames:
        return []

    point_rows = list(workbook[POINT_OF_SALE_SHEET].iter_rows(values_only=True))
    comuna_rows = list(workbook[COMUNA_SHEET].iter_rows(values_only=True))
    comuna_zone = _build_comuna_zone_map(point_rows)
    week_columns = _current_cycle_week_columns(comuna_rows)
    weekly_sales: dict[tuple[str, int], int] = defaultdict(int)
    weekly_communes: dict[tuple[str, int], set[str]] = defaultdict(set)

    for row in comuna_rows[DATA_START_ROW - 1:]:
        comuna = _row_value(row, 2)
        if not comuna:
            continue
        zone = comuna_zone.get(_normalize(comuna), "Sin zona")
        for week, column in week_columns:
            value = _number(row[column - 1] if column - 1 < len(row) else None)
            if value is None:
                continue
            weekly_sales[(zone, week)] += int(value)
            weekly_communes[(zone, week)].add(comuna)

    result = [
        WeeklyZoneSale(
            zone=zone,
            week=week,
            week_label=f"S{week}",
            sales=sales,
            communes=len(weekly_communes[(zone, week)]),
        )
        for (zone, week), sales in weekly_sales.items()
    ]
    workbook.close()
    return sorted(result, key=lambda item: (item.week, item.zone))


def _build_comuna_zone_map(rows: list[tuple[object, ...]]) -> dict[str, str]:
    if len(rows) < HEADER_ROW:
        return {}
    headers = [_normalize_header(value) for value in rows[HEADER_ROW - 1]]
    try:
        comuna_col = headers.index("comuna") + 1
        zone_col = headers.index("ubicacion") + 1
    except ValueError:
        return {}

    zones_by_comuna: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows[DATA_START_ROW - 1:]:
        comuna = _row_value(row, comuna_col)
        zone = _row_value(row, zone_col)
        if comuna and zone:
            zones_by_comuna[_normalize(comuna)][zone] += 1

    return {
        comuna: counter.most_common(1)[0][0]
        for comuna, counter in zones_by_comuna.items()
        if counter
    }


def _current_cycle_week_columns(rows: list[tuple[object, ...]]) -> list[tuple[int, int]]:
    if len(rows) < HEADER_ROW:
        return []
    header = rows[HEADER_ROW - 1]
    week_headers = [
        (column, value)
        for column, value in enumerate(header, start=1)
        if isinstance(value, int)
    ]
    second_week_one_index = _second_week_one_index(week_headers)
    if second_week_one_index is None:
        return []

    columns: list[tuple[int, int]] = []
    for column, value in week_headers[second_week_one_index:]:
        if isinstance(value, int):
            columns.append((value, column))
    return columns


def _second_week_one_index(week_headers: list[tuple[int, int]]) -> int | None:
    week_one_count = 0
    for index, (_, value) in enumerate(week_headers):
        if value == 1:
            week_one_count += 1
            if week_one_count == 2:
                return index
    return None


def _normalize_header(value: object) -> str:
    return _normalize(_text(value) or "")


def _normalize(value: str) -> str:
    return (
        value.strip()
        .lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _row_value(row: tuple[object, ...], column: int) -> str | None:
    return _text(row[column - 1] if column - 1 < len(row) else None)


def _number(value: object) -> int | None:
    if isinstance(value, int | float):
        return int(value)
    return None
