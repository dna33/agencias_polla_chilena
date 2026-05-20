from __future__ import annotations

from openpyxl import Workbook

from app.weekly_sales import parse_weekly_workbook, parse_weekly_workbooks


def test_parse_weekly_sheet_with_header_on_first_row(tmp_path):
    path = tmp_path / "Base Sem 13.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoja1"
    sheet.append([
        "N°",
        "  Lotos",
        "Nombre Agente",
        "Comuna",
        "Est. Com.",
        "Vta.Sem.13",
        "Latitud",
        "Longitud",
        "Ubicación",
    ])
    sheet.append([1, 123456, "Agencia Uno", "SANTIAGO", "Direccion", 1000, "-33.4", "-70.6", "RM Norte"])
    workbook.save(path)

    result = parse_weekly_workbook(path)

    assert result.skipped_sheets == []
    assert len(result.rows) == 1
    assert result.rows[0].week == 13
    assert result.rows[0].lotos_code == "123456"
    assert result.rows[0].weekly_sales == 1000
    assert result.rows[0].territory == "RM Norte"


def test_parse_weekly_sheet_with_header_after_metadata_rows(tmp_path):
    path = tmp_path / "Base Semana 15.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LOTO_ PtoVta"
    sheet.append(["metadata"])
    sheet.append([])
    sheet.append(["metadata"])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append(["F", 1])
    sheet.append(["N°", "  Lotos", "Nombre Agente", "Status", "Vta.Sem.15", "Master"])
    sheet.append([1, 654321, "Agencia Dos", "activo", 2500, 654321])
    workbook.save(path)

    result = parse_weekly_workbook(path)

    assert result.rows[0].source_row == 9
    assert result.rows[0].week == 15
    assert result.rows[0].operational_status == "activo"
    assert result.rows[0].master_code == "654321"


def test_filename_week_overrides_stale_sales_header(tmp_path):
    path = tmp_path / "Base Sem 10.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LOTO_ PtoVta"
    sheet.append(["N°", "  Lotos", "Nombre Agente", "Vta.Sem.09"])
    sheet.append([1, 123456, "Agencia Uno", 3210])
    workbook.save(path)

    result = parse_weekly_workbook(path)

    assert result.rows[0].week == 10
    assert result.rows[0].weekly_sales == 3210


def test_parse_historical_sheet_with_multiple_week_columns(tmp_path):
    path = tmp_path / "Venta Loto Semana 17 vs Anteriores Reporte DAZ vsimple.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LOTO_ PtoVta"
    for _ in range(7):
        sheet.append([])
    sheet.append([
        "N°",
        "  Lotos",
        "Nombre Agente",
        "Comuna",
        "Est. Com.",
        "Vta.Sem.1",
        "Vta.Sem.2",
        "Vta.Sem.3",
        "Latitud",
        "Longitud",
        "Ubicación",
    ])
    sheet.append([1, 123456, "Agencia Uno", "SANTIAGO", "Activo", 1000, 2500, 3100, "-33.4", "-70.6", "RM Norte"])
    workbook.save(path)

    result = parse_weekly_workbook(path)

    assert len(result.rows) == 3
    assert [row.week for row in result.rows] == [1, 2, 3]
    assert [row.weekly_sales for row in result.rows] == [1000, 2500, 3100]


def test_parse_weekly_workbooks_deduplicates_overlapping_weeks_preferring_later_file(tmp_path):
    historical = tmp_path / "Venta Loto Semana 17 vs Anteriores Reporte DAZ vsimple.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LOTO_ PtoVta"
    sheet.append(["N°", "  Lotos", "Nombre Agente", "Vta.Sem.17"])
    sheet.append([1, 123456, "Agencia Uno", 1000])
    workbook.save(historical)

    incremental = tmp_path / "Base Semana 18.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LOTO_ PtoVta"
    sheet.append(["N°", "  Lotos", "Nombre Agente", "Vta.Sem.17", "Vta.Sem.18"])
    sheet.append([1, 123456, "Agencia Uno", 1200, 1500])
    workbook.save(incremental)

    rows = parse_weekly_workbooks([historical, incremental])

    assert [(row.week, row.weekly_sales) for row in rows] == [(17, 1200), (18, 1500)]
