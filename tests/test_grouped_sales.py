from __future__ import annotations

from openpyxl import Workbook

from app.grouped_sales import parse_weekly_zone_evolution


def test_parse_weekly_zone_evolution_from_grouped_workbook(tmp_path):
    path = tmp_path / "Base Semana 15.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)

    comuna_sheet = workbook.create_sheet("LOTO_ Comuna")
    point_sheet = workbook.create_sheet("LOTO_ PtoVta")

    for _ in range(7):
        comuna_sheet.append([])
    comuna_sheet.append(["Reg.", "Comuna", "Ejecutivo", 1, 2, 3, 1, 2, 3, 4, 5, 6])
    comuna_sheet.append([13, "SANTIAGO", "RENATO", 10, 20, 30, 100, 200, 300, 400, 500, 600])
    comuna_sheet.append([5, "VALPARAISO", "RODRIGO", 10, 20, 30, 1000, 2000, 3000, 4000, 5000, 6000])

    for _ in range(7):
        point_sheet.append([])
    point_sheet.append(["N°", "  Lotos", "Nombre Agente", "Comuna", "Ubicación"])
    point_sheet.append([1, 111111, "Agencia Uno", "SANTIAGO", "RM Norte"])
    point_sheet.append([2, 222222, "Agencia Dos", "VALPARAISO", "Norte"])
    workbook.save(path)

    rows = parse_weekly_zone_evolution(path)

    week_1 = [row for row in rows if row.week == 1]
    week_6 = [row for row in rows if row.week == 6]
    assert {row.zone for row in week_1} == {"RM Norte", "Norte"}
    assert next(row for row in week_1 if row.zone == "RM Norte").sales == 100
    assert next(row for row in week_1 if row.zone == "Norte").sales == 1000
    assert next(row for row in week_6 if row.zone == "RM Norte").sales == 600
